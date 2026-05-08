"""
Merge the Goldenray Energy NZ master catalogue (File 1) with the pricing-tool
Materials sheet (File 2) into a single cleaned products_merged.xlsx.

This is the Phase 0 step before importing into the products DB table.

Usage:
    python scripts/merge_product_excels.py <master.xlsx> <pricing.xlsx> <output.xlsx>

What it does:
  1. Loads File 1 — finds the real header row, skips instructional rows.
  2. Loads File 2's Materials sheet — extracts cost + margin% per item.
  3. Joins by SKU (Product ID); for File-2-only rows it falls back to a
     normalised-name match against File 1.
  4. Fills missing margin% with 30% (the modal value in File 2).
  5. Computes sell_excl_gst and sell_incl_gst from cost + margin + 15% GST.
  6. Parses freeform availability_notes like "Due 15 Jun 2026" into
     a proper DATE in the available_from column.
  7. Parses wattage from product names ("595W") into wattage_w.
  8. Drops File 1's instructional/template footer rows.
  9. Writes products_merged.xlsx with a clean schema and prints a summary.

Re-runnable: same input always produces the same output. Safe to run on
revised source files; treat the output as the single authoritative file
to feed into the importer.
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

GST_RATE = 0.15
DEFAULT_MARGIN_PCT = 30.0  # File 2's modal value, used when missing

# Output column order for products_merged.xlsx
OUTPUT_COLUMNS = [
    "sku", "category", "subcategory", "brand", "name", "description",
    "cost_nzd", "default_margin_pct", "sell_excl_gst", "sell_incl_gst",
    "unit", "stock_status", "qty_available", "moq",
    "availability_notes", "available_from",
    "website_category", "wattage_w", "source", "needs_review",
]


@dataclass
class ProductRow:
    sku: str | None = None
    category: str | None = None
    subcategory: str | None = None
    brand: str | None = None
    name: str | None = None
    description: str | None = None
    cost_nzd: float | None = None
    default_margin_pct: float | None = None
    unit: str | None = None
    stock_status: str | None = None
    qty_available: int | None = None
    moq: int | None = None
    availability_notes: str | None = None
    available_from: date | None = None
    website_category: str | None = None
    wattage_w: int | None = None
    source: str = "file1"
    needs_review_reasons: list[str] = field(default_factory=list)

    @property
    def sell_excl_gst(self) -> float | None:
        if self.cost_nzd is None or self.default_margin_pct is None:
            return None
        return round(self.cost_nzd * (1 + self.default_margin_pct / 100), 2)

    @property
    def sell_incl_gst(self) -> float | None:
        excl = self.sell_excl_gst
        return round(excl * (1 + GST_RATE), 2) if excl is not None else None

    @property
    def needs_review(self) -> str:
        return "; ".join(self.needs_review_reasons) if self.needs_review_reasons else ""

    def as_row(self) -> list[Any]:
        return [
            self.sku, self.category, self.subcategory, self.brand,
            self.name, self.description,
            self.cost_nzd, self.default_margin_pct,
            self.sell_excl_gst, self.sell_incl_gst,
            self.unit, self.stock_status, self.qty_available, self.moq,
            self.availability_notes,
            self.available_from.isoformat() if self.available_from else None,
            self.website_category, self.wattage_w,
            self.source, self.needs_review,
        ]


# ── Parsing helpers ──────────────────────────────────────────────────────────

WATTAGE_RE = re.compile(r"\b(\d{2,4})\s*W\b", re.IGNORECASE)
DATE_RE = re.compile(
    r"\b(\d{1,2})\s+"
    r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+"
    r"(20\d{2})\b",
    re.IGNORECASE,
)
MONTH = {m: i for i, m in enumerate(
    "jan feb mar apr may jun jul aug sep oct nov dec".split(), start=1)}


def parse_wattage(name: str | None) -> int | None:
    """Pull a wattage out of a product name like 'Phono 595W Draco' → 595."""
    if not name:
        return None
    m = WATTAGE_RE.search(name)
    if not m:
        return None
    val = int(m.group(1))
    # Solar panel / inverter wattages are 100–800. Reject suspiciously
    # huge numbers (could be model codes that happened to end with 'W').
    if 50 <= val <= 1500:
        return val
    return None


def parse_available_from(notes: str | None) -> date | None:
    """Extract a DATE from notes like 'Due 15 Jun 2026' → 2026-06-15."""
    if not notes:
        return None
    m = DATE_RE.search(notes)
    if not m:
        return None
    day, mon, year = int(m.group(1)), MONTH[m.group(2).lower()[:3]], int(m.group(3))
    try:
        return date(year, mon, day)
    except ValueError:
        return None


def normalise_name(s: str | None) -> str:
    """For fuzzy name-based joining when SKUs aren't shared between files."""
    if not s:
        return ""
    s = s.lower()
    # collapse whitespace, drop punctuation/dashes
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    # remove brand prefix patterns common to File 2 ("Phono Solar-Phono Solar 595W")
    return s


def to_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


# ── File 1 loader ────────────────────────────────────────────────────────────

def load_master(path: Path) -> list[ProductRow]:
    """Read File 1's master template, skipping documentation rows."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # The real header sits on the row literally labelled "Category", "Sub Category", ...
    header_idx = next(
        (i for i, r in enumerate(rows)
         if r and r[0] == "Category" and r[1] == "Sub Category"),
        None,
    )
    if header_idx is None:
        sys.exit("Could not find File 1 header row (expected 'Category', 'Sub Category', ...)")

    products: list[ProductRow] = []
    skipped_instructional = 0

    for r in rows[header_idx + 1:]:
        if not r or all(c is None for c in r):
            continue
        # Pad to expected width
        r = list(r) + [None] * (13 - len(r))
        cat, sub, brand, pid, name, desc, price, unit, stock, qty, moq, notes, web_cat = r[:13]

        # Skip rows that are template/footer instructions:
        #   - Product ID is None or non-numeric
        #   - Brand is "Margin $", "Generic" without a real product
        #   - Category is "Purchase Price"
        if cat in ("Purchase Price",):
            skipped_instructional += 1
            continue
        if pid is None:
            skipped_instructional += 1
            continue
        # Numeric SKU check
        try:
            int(str(pid).strip())
        except (ValueError, TypeError):
            skipped_instructional += 1
            continue

        p = ProductRow(
            sku=str(pid).strip(),
            category=cat,
            subcategory=sub,
            brand=brand,
            name=name,
            description=desc,
            cost_nzd=to_float(price),
            unit=unit,
            stock_status=normalise_stock_status(stock),
            qty_available=to_int(qty),
            moq=to_int(moq),
            availability_notes=notes,
            website_category=web_cat,
            source="file1",
        )
        p.available_from = parse_available_from(p.availability_notes)
        p.wattage_w = parse_wattage(p.name)

        # Track data quality issues
        if p.cost_nzd is None:
            p.needs_review_reasons.append("missing_cost")
        if not p.brand:
            p.needs_review_reasons.append("missing_brand")
        if not p.category:
            p.needs_review_reasons.append("missing_category")

        products.append(p)

    print(f"  File 1: {len(products)} products loaded (skipped {skipped_instructional} instructional rows)")
    return products


def normalise_stock_status(s: Any) -> str | None:
    if not s:
        return None
    sv = str(s).strip().lower()
    if "in stock" in sv:
        return "in_stock"
    if "backorder" in sv or "back order" in sv:
        return "backorder"
    if "discontinue" in sv:
        return "discontinued"
    return sv  # leave as-is, flag below


# ── File 2 loader ────────────────────────────────────────────────────────────

def load_pricing(path: Path) -> list[dict]:
    """
    Read File 2's Materials sheet. Each row may or may not have margin% filled.
    Returns a list of dicts with description, cost_nzd, margin_pct.
    File 2 has no SKUs of its own — we'll match to File 1 by name where possible.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Materials" not in wb.sheetnames:
        wb.close()
        sys.exit("File 2 has no 'Materials' sheet — wrong file?")
    ws = wb["Materials"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    items: list[dict] = []
    for r in rows[1:]:  # skip header
        if not r or r[1] is None:
            continue
        desc = str(r[1]).strip()
        cost = to_float(r[3])
        margin = to_float(r[4])
        if cost is None:
            continue
        items.append({
            "section": r[0],          # category marker like "A. PV Modules"
            "description": desc,
            "cost_nzd": cost,
            "margin_pct": margin,
            "norm": normalise_name(desc),
        })
    print(f"  File 2: {len(items)} material rows loaded")
    return items


# ── Merge ────────────────────────────────────────────────────────────────────

def merge(file1: list[ProductRow], file2: list[dict]) -> list[ProductRow]:
    """Merge File 2's margin info into File 1's products by name match.
       Add File-2-only items as separate ProductRow entries (no SKU)."""
    f1_by_norm: dict[str, ProductRow] = {}
    for p in file1:
        if p.name:
            f1_by_norm[normalise_name(p.name)] = p

    matched, file1_with_margin = 0, 0
    file2_only: list[ProductRow] = []

    for item in file2:
        # Try fuzzy match: take significant chunk of normalised description
        norm_desc = item["norm"]
        # Try full-string match first
        match = f1_by_norm.get(norm_desc)
        if not match:
            # Try matching by 25-character chunk
            for n, p in f1_by_norm.items():
                if len(n) > 20 and (norm_desc in n or n in norm_desc):
                    match = p
                    break

        if match:
            matched += 1
            # File 1's price IS the cost (confirmed equal in earlier analysis).
            # If File 2 has a margin% for this item, use it.
            if item["margin_pct"] is not None and match.default_margin_pct is None:
                match.default_margin_pct = item["margin_pct"]
                file1_with_margin += 1
        else:
            # File-2-only line item — create a stub product, mark for review
            stub = ProductRow(
                sku=None,
                name=item["description"],
                description=item["description"],
                category=item["section"],  # e.g. "A. PV Modules"
                cost_nzd=item["cost_nzd"],
                default_margin_pct=item["margin_pct"],
                unit="EA",
                stock_status="unknown",
                source="file2",
            )
            stub.wattage_w = parse_wattage(stub.name)
            stub.needs_review_reasons.append("file2_only_no_sku")
            file2_only.append(stub)

    # Backfill default margin where still missing
    backfilled = 0
    for p in file1 + file2_only:
        if p.default_margin_pct is None:
            p.default_margin_pct = DEFAULT_MARGIN_PCT
            p.needs_review_reasons.append("margin_defaulted")
            backfilled += 1

    print(f"  Match: {matched} of {len(file2)} File 2 rows matched a File 1 product")
    print(f"         (File 1 products that gained a margin from File 2: {file1_with_margin})")
    print(f"         {len(file2_only)} File-2-only rows kept as un-SKU'd stubs")
    print(f"  Defaulted margin% to {DEFAULT_MARGIN_PCT}% on {backfilled} rows")

    return file1 + file2_only


# ── Output ───────────────────────────────────────────────────────────────────

def write_output(products: list[ProductRow], out: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(OUTPUT_COLUMNS)
    for p in products:
        ws.append(p.as_row())
    # Auto-size-ish: sample first row widths
    for col_idx, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, len(col) + 2)
    wb.save(out)


def summarise(products: list[ProductRow]) -> None:
    print()
    print(f"OUTPUT SUMMARY ({len(products)} products):")
    cats = Counter(p.category for p in products if p.category)
    brands = Counter(p.brand for p in products if p.brand)
    sources = Counter(p.source for p in products)
    review = sum(1 for p in products if p.needs_review_reasons)
    has_wattage = sum(1 for p in products if p.wattage_w)
    has_avail_date = sum(1 for p in products if p.available_from)
    has_cost = sum(1 for p in products if p.cost_nzd is not None)

    print(f"  By source:    {dict(sources)}")
    print(f"  Categories:   {len(cats)} distinct")
    for c, n in cats.most_common(8):
        print(f"    {n:4d}  {c}")
    print(f"  Brands:       {len(brands)} distinct (top 5):")
    for b, n in brands.most_common(5):
        print(f"    {n:4d}  {b}")
    print(f"  With cost:                    {has_cost} / {len(products)}")
    print(f"  With wattage parsed:          {has_wattage}")
    print(f"  With backorder date parsed:   {has_avail_date}")
    print(f"  Flagged for human review:     {review}  ({review/len(products)*100:.1f}%)")
    if review:
        reasons = Counter()
        for p in products:
            for r in p.needs_review_reasons:
                reasons[r] += 1
        print("    Review reasons:")
        for r, n in reasons.most_common():
            print(f"      {n:4d}  {r}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("master", type=Path, help="File 1 — master catalogue Excel")
    ap.add_argument("pricing", type=Path, help="File 2 — pricing tool Excel")
    ap.add_argument("output", type=Path, help="Output products_merged.xlsx")
    args = ap.parse_args()

    for p, label in [(args.master, "master"), (args.pricing, "pricing")]:
        if not p.exists():
            sys.exit(f"{label} file not found: {p}")

    print(f"Loading master:  {args.master}")
    file1 = load_master(args.master)
    print(f"Loading pricing: {args.pricing}")
    file2 = load_pricing(args.pricing)
    print()
    print("Merging:")
    merged = merge(file1, file2)

    write_output(merged, args.output)
    print(f"\nWrote: {args.output}")
    summarise(merged)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
